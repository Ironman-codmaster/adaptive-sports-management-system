# import pandas as pd
# import os
# import shutil
# from tkinter import Label, filedialog
# from openpyxl import Workbook, load_workbook
#
#
# destination_folder = "C:/"
# os.makedirs(destination_folder, exist_ok=True)
# file_path = filedialog.askopenfilename(
#     title="Fotosuratni tanlang",
#     filetypes=[("Faqat Excel fayl", "*.xlsx")]
# )
#
# # Загружаем файл
# wb = load_workbook(file_path)
# ws = wb.active  # Или ws = wb['Sheet1']
#
# # Пример: вывести все строки
# for row in ws.iter_rows(values_only=True):
#     print(row)
#
# # Доступ к ячейке A1
# print(ws['A1'].value)

#------------------------------------------------------------------------

from openpyxl import Workbook
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Side, Border
import os


# Создание книги и листа
wb = Workbook()
ws = wb.active
ws.sheet_view.zoomScale = 130  # можно задать 80, 120 и т.д.

ws.title = "Баённома"
time_n_r12 = Font(name='Times New Roman', size=12, bold=True)


# Настройка ширины столбцов для вида таблицы
ws.column_dimensions['A'].width = 3.14
ws.column_dimensions['B'].width = 28
ws.column_dimensions['C'].width = 13.29
ws.column_dimensions['D'].width = 8.86
ws.column_dimensions['E'].width = 15.57
ws.column_dimensions['F'].width = 9.43
ws.column_dimensions['G'].width = 8.43

# Вставка изображений
img_left = Image("test_images/logoparauz1.png")  # Левая картинка (вы прислали только одну, её используем дважды)
img_right = Image("test_images/laylak.jpg")  # Правая картинка (в реальности должна быть другая)

# Изменим размер изображения (под ячейки)
img_left.width = 90
img_left.height = 100
img_right.width = 80
img_right.height = 130


# Заголовок (объединённые ячейки)
ws.merge_cells('A1:G1')

# Вставляем в нужные координаты
ws.add_image(img_left, "A1")   # Левый верхний угол
ws.add_image(img_right, "G1")  # Правый верхний угол

ws['A1'] = "Пара енгил атлетика бўйича Ўзбекистон чемпионати\nT/46-47 тоифасида 100 метр масофага\nюгуриш мусобақа\nБАЁННОМАСИ"
ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
#ws['A1'].font = Font(bold=True)
ws.row_dimensions[1].height = 90
ws['A1'].font = time_n_r12

# Женская категория
ws.merge_cells('A2:G2')
ws['A2'] = "АЁЛЛАР"
ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
#ws['A2'].font = Font(bold=True)
ws['A2'].font = time_n_r12
ws.row_dimensions[2].height = 18


# Дата и город
ws.merge_cells('A3:B3')
ws['A3'] = "6-8 май 2024 йил"
ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
ws['A3'].font = time_n_r12

ws.merge_cells('E3:G3')
ws['E3'] = "Ангрен шаҳри"
ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
ws['E3'].font = time_n_r12

# ws['B3'].alignment = ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
# ws['B3'].font = ws['E3'].font = Font(bold=True)

# Заголовки таблицы
headers = ['№', 'Ф.И.О', 'Туғилган йили', 'Вилоят', 'Кўкрак рақами', 'Натижа', 'Ўрин']
thick_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

for col, header in enumerate(headers, start=1):
    cell = ws.cell(row=4, column=col)
    cell.value = header
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    #cell.font = Font(bold=True)
    cell.font = time_n_r12
    cell.border = thick_border

# Сохраняем файл
excel_fayl_nomi = "bayonnoma.xlsx"
try:
    wb.save(excel_fayl_nomi)
    os.startfile(excel_fayl_nomi)
except Exception as e:
      print("Bayonnoma fayli excelda ochiq holatda. Yoki boshqa hatolik yuz berdi.\nHatolik: "+str(e))
