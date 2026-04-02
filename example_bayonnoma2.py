import shutil

from openpyxl import Workbook
import docx
from docx.shared import Pt
from docx.shared import Inches
import os
import subprocess

from docx import Document
from docx.shared import Cm, Inches
from docx.enum.table import WD_ALIGN_VERTICAL
from docx.enum.text import WD_PARAGRAPH_ALIGNMENT


from openpyxl.drawing.image import Image
#from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Font, Side, Border


import sys
from tkinter import filedialog
from tkinter import messagebox as mb
from openpyxl import load_workbook
from tkinter import ttk

from todox import *
import pandas as pd


def exeldanolish():
    destination_folder = "C:/"
    os.makedirs(destination_folder, exist_ok=True)
    file_path = filedialog.askopenfilename(
        title="Exel faylini tanlang",
        filetypes=[("Fayl turi", "*.xlsx *.xls")]
    )
    # Загружаем файл
    wb = load_workbook(file_path)
    ws = wb.active  # Или ws = wb['Sheet1']

    # Пример: вывести все строки
    # for row in ws.iter_rows(min_col=1, max_col=9, values_only=True):
    #     print(row)

    result = ws.iter_rows(min_col=1, max_col=9, values_only=True)
    isht_list = []
    for record in result:
        isht_list.append(record)



    # Преобразуем твои данные в список словарей (исключаем заголовки)
    raw_data = isht_list

    columns = ['№', 'Ф.И.О', 'Тугилганили', 'Жинси', 'Класс', '1-тур', '2-тур', '3-тур', 'Вилоят']

    # Создаем DataFrame
    df = pd.DataFrame(raw_data, columns=columns)

    # Создаем Excel-файл
    with pd.ExcelWriter("спортсмены.xlsx", engine='openpyxl') as writer:
        # Группируем по Класс и Жинси
        grouped = df.groupby(['Класс', 'Жинси'])

        for (klass, gender), group in grouped:
            # Фильтруем только нужные строки
            filtered = group[
                (group['1-тур'] == '100м') | (group['2-тур'] == '100м') | (group['3-тур'] == '100м')
                ]

            # Берем нужные столбцы
            result = filtered[['№', 'Ф.И.О', 'Тугилганили', 'Вилоят']]

            # Название листа: например, "T-12_Э"
            sheet_name = f"{klass}_{gender}"

            # Сохраняем на отдельный лист
            result.to_excel(writer, sheet_name=sheet_name, index=False)


def to_bayonnoma(self, l1nomi, l2bayon):
    # Создание книги и листа
    wb = Workbook()
    ws = wb.active
    ws.sheet_view.zoomScale = 130  # можно задать 80, 120 и т.д.

    ws.title = "Баённома"
    time_n_r12 = Font(name='Times New Roman', size=12, bold=True)

    # Настройка ширины столбцов для вида таблицы
    ws.column_dimensions['A'].width = 3.14
    ws.column_dimensions['B'].width = 28
    ws.column_dimensions['C'].width = 13.29
    ws.column_dimensions['D'].width = 8.86
    ws.column_dimensions['E'].width = 15.57
    ws.column_dimensions['F'].width = 9.43
    ws.column_dimensions['G'].width = 8.43

    # Вставка изображений
    img_left = Image(
        "test_images/logoparauz1.png")  # Левая картинка (вы прислали только одну, её используем дважды)
    img_right = Image("test_images/laylak.jpg")  # Правая картинка (в реальности должна быть другая)

    # Изменим размер изображения (под ячейки)
    img_left.width = 90
    img_left.height = 100
    img_right.width = 80
    img_right.height = 130

    # Заголовок (объединённые ячейки)
    ws.merge_cells('A1:G1')

    # Вставляем в нужные координаты
    ws.add_image(img_left, "A1")  # Левый верхний угол
    ws.add_image(img_right, "G1")  # Правый верхний угол

    if l1nomi[1] == "1-Klassni tanlang":
        l1nomi[1] = ""
    else:
        l1nomi[1] = l1nomi[1] + " "
    if l1nomi[2] == "2-Klassni tanlang":
        l1nomi[2] = ""
    else:
        l1nomi[2] = l1nomi[2] + " "
    if l1nomi[3] == "3-Klassni tanlang":
        l1nomi[3] = ""
    # print(l1nomi[4])
    if l1nomi[4] == "100м":
        l1nomi[4] = "100 метр масофага\nюгуриш бўйича"
    if l1nomi[4] == "200м":
        l1nomi[4] = "200 метр масофага\nюгуриш бўйича"
    if l1nomi[4] == "400м":
        l1nomi[4] = "400 метр масофага\nюгуриш бўйича"
    if l1nomi[4] == "1500м":
        l1nomi[4] = "1500 метр масофага\nюгуриш бўйича"
    if l1nomi[4] == "Ядро":
        l1nomi[4] = "Ядро улоқтириш\nбўйича"
    if l1nomi[4] == "Диск":
        l1nomi[4] = "Диск улоқтириш\nбўйича"
    if l1nomi[4] == "Клаб":
        l1nomi[4] = "CLUB THROW улоқтириш\nбўйича"
    if l1nomi[4] == "Узунлик":
        l1nomi[4] = "Узунликка сакраш\nбўйича"
    if l1nomi[4] == "Найза":
        l1nomi[4] = "Ядро улоқтириш\nбўйича"
    if l1nomi[4] == "Баландлик":
        l1nomi[4] = "Баландликка сакраш\nбўйича"
    if l1nomi[4] == "Арава":
        l1nomi[4] = "100 метр масофага\nюгуриш бўйича"

    ws[
        'A1'] = f"Пара енгил атлетика бўйича Ўзбекистон чемпионати\n{l1nomi[1]}{l1nomi[2]}{l1nomi[3]} тоифасида {l1nomi[4]} мусобақа\nБАЁННОМАСИ"
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    # ws['A1'].font = Font(bold=True)
    ws.row_dimensions[1].height = 90
    ws['A1'].font = time_n_r12
    if l1nomi[0] == "Э":
        jns = "ЭРКАКЛАР"
    else:
        jns = "АЁЛЛАР"
    # Женская категория
    ws.merge_cells('A2:G2')
    ws['A2'] = jns
    ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
    # ws['A2'].font = Font(bold=True)
    ws['A2'].font = time_n_r12
    ws.row_dimensions[2].height = 18

    # Дата и город
    ws.merge_cells('A3:B3')
    ws['A3'] = "6-8 май 2024 йил"
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['A3'].font = time_n_r12

    ws.merge_cells('E3:G3')
    ws['E3'] = "Ангрен шаҳри"
    ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
    ws['E3'].font = time_n_r12

    # ws['B3'].alignment = ws['E3'].alignment = Alignment(horizontal="center", vertical="center")
    # ws['B3'].font = ws['E3'].font = Font(bold=True)

    # Заголовки таблицы
    headers = ['№', 'Ф.И.О', 'Туғилган йили', 'Вилоят', 'Кўкрак рақами', 'Натижа', 'Ўрин']
    thick_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'))

    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=4, column=col)
        cell.value = header
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        # cell.font = Font(bold=True)
        cell.font = time_n_r12
        cell.border = thick_border

        # Telo таблицы
        thick_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin'))
        # Listni hujjatga moslashtirib olamiz

        bay_uch = [item[:3] + (item[8],) for item in l2bayon]

        # Пишем таблицу, начиная с 6-й строки (row + 5)
        for row_idx, row_data in enumerate(bay_uch, start=1):
            for col_idx, cell_value in enumerate(row_data, start=1):
                cell = ws.cell(row=row_idx + 4, column=col_idx)
                cell.value = cell_value
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.font = Font(name='Times New Roman', size=12, bold=False)
                cell.border = thick_border
            for col_idx, cell_value in enumerate(["", "", ""], start=1):
                cell = ws.cell(row=row_idx + 4, column=col_idx + 4)
                cell.value = cell_value
                cell.border = thick_border

    # Сохраняем файл
    # excel_fayl_nomi = "bayonnoma.xlsx"
    # excel_fayl_nomi = (l1nomi[1]+l1nomi[2]+l1nomi[3]+l1nomi[4][:5]+"_bayonnoma.xlsx").replace(" ",'_')

    # Получаем путь к папке "Документы"
    documents_folder = os.path.join(os.path.expanduser("~"), "Documents")

    # Формируем полный путь к файлу
    excel_fayl_nomi = os.path.join(documents_folder, "bayonnoma.xlsx")

    try:
        wb.save(excel_fayl_nomi)
        os.startfile(excel_fayl_nomi)
    except Exception as e:
        ex = "Bayonnoma fayli excelda ochiq holatda. Yoki boshqa hatolik yuz berdi.\nHatolik: " + str(e)
        return ex

exeldanolish()
