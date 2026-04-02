import customtkinter
import sys
from tkinter import filedialog
from tkinter import messagebox as mb
from openpyxl import load_workbook
from tkinter import ttk

from todox import *

from bayonnoma_umumiy import Umumiy_Bayon
class Protokol:

    isht_list = []
    def protokol(self):
        # Функция для безопасной загрузки ресурсов (работает и в .py, и в .exe)
        def resource_path(relative_path):
            try:
                base_path = sys._MEIPASS  # путь при запуске .exe
            except AttributeError:
                base_path = os.path.abspath(".")  # путь при запуске .py
            return os.path.join(base_path, relative_path)

        # Путь к папке с изображениями
        image_path = resource_path("test_images")

        self.second_frame = customtkinter.CTkFrame(self, corner_radius=0, fg_color="transparent")
        self.second_frame.grid_columnconfigure(0, weight=1)

        # Treevew O'quvchilarni ko'rsatib turuvchi oyna

        self.tree_oqu_frame = customtkinter.CTkFrame(self.second_frame, corner_radius=0, fg_color="transparent",
                                                     width=780, height=610, border_width=1,
                                                     border_color=("black", "white"))
        self.tree_oqu_frame.place(x=5, y=5)

        # Scrollar
        scrollbarx = ttk.Scrollbar(self.tree_oqu_frame, orient="horizontal")
        scrollbary = ttk.Scrollbar(self.tree_oqu_frame, orient="vertical")

        # Trevewning ko'rinishi
        style = ttk.Style()
        style.configure("mystyle.Treeview", highlightthickness=0, bd=0, font=('Calibri', 10),
                        rowheght=25)  # Modify the font of the body
        style.configure("mystyle.Treeview.Heading", font=('Calibri', 12, 'bold'))  # Modify the font of the headings
        style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])  # Remove the borders

        # Treevew
        my_tree = ttk.Treeview(self.tree_oqu_frame, style="mystyle.Treeview")
        my_tree.place(x=5, y=5, width=750, height=580)
        my_tree.configure(yscrollcommand=scrollbary.set, xscrollcommand=scrollbarx.set)
        my_tree.configure(selectmode="extended")
        # scrollbar larning korinishi va joylashgan joyi
        scrollbary.configure(command=my_tree.yview)
        scrollbarx.configure(command=my_tree.xview)

        scrollbary.place(x=755, y=5, width=15, height=580)
        scrollbarx.place(x=5, y=586, width=750, height=15)

        my_tree.configure(
            columns=(
                "IN",
                "FIO",
                "Tugulganyili",
                "Jinsi",
                "Klass",
                "1tur",
                "2tur",
                "3tur",
                "Viloyati"
            )
        )

        # Headings
        my_tree.heading("#0", text="N")
        my_tree.heading("IN", text="ExN")
        my_tree.heading("FIO", text="F.I.O")
        my_tree.heading("Tugulganyili", text="TugYili")
        my_tree.heading("Jinsi", text="Jinsi")
        my_tree.heading("Klass", text="Klass")
        my_tree.heading("1tur", text="1-tur")
        my_tree.heading("2tur", text="2-tur")
        my_tree.heading("3tur", text="3-tur")
        my_tree.heading("Viloyati", text="Viloyati")

        # Colums
        my_tree.column("#0", minwidth=0, width=40)  # N
        my_tree.column("#1", minwidth=0, width=40)  # ExN
        my_tree.column("#2", minwidth=0, width=150)  # FIO
        my_tree.column("#3", minwidth=0, width=50)  # Tug'ulgan yili
        my_tree.column("#4", minwidth=0, width=50)  # Jinsi
        my_tree.column("#5", minwidth=0, width=50)  # Klass
        my_tree.column("#6", minwidth=0, width=50)  # 1-tur
        my_tree.column("#7", minwidth=0, width=50)  # 2-tur
        my_tree.column("#8", minwidth=0, width=50)  # 3-tur
        my_tree.column("#9", minwidth=0, width=100)  # viloyati

        self.exel_buttons_frame = customtkinter.CTkFrame(self.second_frame, corner_radius=0, fg_color="transparent",
                                                       border_width=1,
                                                       border_color=("black", "white"), width=125, height=600)
        self.exel_buttons_frame.place(x=800, y=5)

        def exeldanolish():
            destination_folder = "C:/"
            os.makedirs(destination_folder, exist_ok=True)
            file_path = filedialog.askopenfilename(
                title="Exel faylini tanlang",
                filetypes=[("Fayl turi", "*.xlsx *.xls")]
            )
            # Загружаем файл
            wb = load_workbook(file_path)
            ws = wb.active  # Или ws = wb['Sheet1']

            # Пример: вывести все строки
            # for row in ws.iter_rows(min_col=1, max_col=9, values_only=True):
            #     print(row)

            result = ws.iter_rows(min_col=1, max_col=9, values_only=True)
            #Treni tozalab olish
            for row in my_tree.get_children():
                my_tree.delete(row)

            # o'quvchilar treevewda zebra bo'lib kornshi
            tr = 1
            my_tree.tag_configure('oddrow', background="white")
            my_tree.tag_configure('evenrow', background="#ecf8fe")
            #result = (1, 2, 3, 4, 5, 6, 7, 8, 9)
            # treewevga ma'lumotlarni kiritish
            Protokol.isht_list = []
            for record in result:
                Protokol.isht_list.append(record)

            to_tree = Protokol.isht_list[1:]
            # print(to_tree)

            for record in to_tree:
                if tr % 2 == 0:
                    my_tree.insert(parent='', index='end', text=str(tr), values=(
                        record[0], record[1], record[2], record[3], record[4], record[5], record[6], record[7], record[8]),
                                   tags=('evenrow',))
                else:
                    my_tree.insert(parent='', index='end', text=str(tr), values=(
                        record[0], record[1], record[2], record[3], record[4],record[5], record[6], record[7], record[8]),
                                   tags=('oddrow',))
                tr += 1

            # 🔄 Обновляем список классов
            r1 = sorted(set(str(ifi[4]).strip() for ifi in to_tree if str(ifi[4]).strip()[0]=="F" or str(ifi[4]).strip()[0]=="T"))
            #r1 = r1.append(" ")
            self.clas1.configure(values=r1)
            self.clas2.configure(values=r1)
            self.clas3.configure(values=r1)
            #print(r1)

        exeldan = customtkinter.CTkButton(self.exel_buttons_frame, text="Exeldan yuklash", image=self.icon_image_excell, width=35, height=35, fg_color="white", hover_color="silver",command=exeldanolish, anchor="center",compound="top")
        exeldan.place(x=10, y=10)

        self.jins = customtkinter.CTkComboBox(
            self.exel_buttons_frame, width=100, border_width=1, state="readonly",
            values=["Э","А"])
        self.jins.place(x=10, y=90)

        r1 = []

        self.clas1 = customtkinter.CTkComboBox(
            self.exel_buttons_frame, width=100, border_width=1, state="readonly",
            values=r1)
        self.clas1.place(x=10, y=130)
        self.clas1.set("1-Klassni tanlang")

        self.clas2 = customtkinter.CTkComboBox(
            self.exel_buttons_frame, width=100, border_width=1, state="readonly",
            values=r1)
        self.clas2.place(x=10, y=170)
        self.clas2.set("2-Klassni tanlang")

        self.clas3 = customtkinter.CTkComboBox(
            self.exel_buttons_frame, width=100, border_width=1, state="readonly",
            values=r1)
        self.clas3.place(x=10, y=210)
        self.clas3.set("3-Klassni tanlang")

        r2=[]
        self.tur = customtkinter.CTkComboBox(
            self.exel_buttons_frame, width=100, border_width=1, state="readonly",
            values=['100м','200м','400м','1500м','Ядро','Диск','Клаб','Узунлик','Найза','Баландлик','Арава'])
        self.tur.place(x=10, y=250)
        self.tur.set("Sport turi")

        def chiqarish():
            pr = Protokol.isht_list
            #print(pr[3])
            jinsi = self.jins.get()
            clas1 = self.clas1.get()
            clas2 = self.clas2.get()
            clas3 = self.clas3.get()
            tur = self.tur.get()

            todoxtitle = [jinsi,clas1,clas2,clas3,tur]

            to_todox = []
            for cls in pr:
                if cls[3] == jinsi:
                     if cls[4] == clas1 or cls[4] == clas2 or cls[4] == clas3:
                         if cls[5] == tur or cls[6] == tur or cls[7] == tur:
                            to_todox.append(cls)

            #print(to_todox)

            docs = Docs()
            ex = docs.to_bayonnoma(todoxtitle, to_todox)
            if ex:
                mb.showerror("Xatolik", ex)

        exeldan = customtkinter.CTkButton(self.exel_buttons_frame, text="Bayonnoma", image=self.icon_image_excell,
                                          width=80, height=35, fg_color="white", hover_color="silver",
                                          command=chiqarish, anchor="center", compound="top")
        exeldan.place(x=10, y=300)


        def chiqarish_umumiy():
            sport = '"1500м"'
            # sport=input("Sport turini kiriting: ")
            sport_turlari_1jadval = ('100м', '1500м', '200м', '400м', 'Арава')
            sport_turlari_2jadval = ('Ядро', 'Баландлик', 'Диск', 'Клаб', 'Найза', 'Узунлик')
            # exeldan olib oddiy jadval hosil qilish
            #Umumiy_Bayon.exeldanolish("1500м")
            if sport in sport_turlari_1jadval:
                # obrabotka_1_jadval
                Umumiy_Bayon.obrabotka(sport)
            elif sport in sport_turlari_2jadval:
                # obrabotka_2_jadval
                Umumiy_Bayon.obrabotka_2jadval(sport)



        exeldan_u = customtkinter.CTkButton(self.exel_buttons_frame, text="Bayonnoma",
                                          width=80, height=35, fg_color="white", hover_color="silver",
                                          command=chiqarish_umumiy, anchor="center", compound="top")
        exeldan_u.place(x=10, y=500)




