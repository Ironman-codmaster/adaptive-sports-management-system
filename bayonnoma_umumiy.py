import shutil
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, Side, Border
from openpyxl.drawing.image import Image
import os
import pandas as pd
from tkinter import filedialog
from tkinter import messagebox as mb
import re

class Umumiy_Bayon:

    def sanitize_sheet_name(self,name):
        return re.sub(r"[\\/?*\[\]:]", "_", name)[:31]  # Excel ограничивает имя листа 31 символом

    def exeldanolish(self,sport_turi):
        destination_folder = "C:/"
        os.makedirs(destination_folder, exist_ok=True)

        file_path = filedialog.askopenfilename(
            title="Exel faylini tanlang",
            filetypes=[("Fayl turi", "*.xlsx *.xls")]
        )
        if not file_path:
            mb.showwarning("Ogohlantirish", "Файл танланмади!")
            return

        wb = load_workbook(file_path)
        ws = wb.active

        result = list(ws.iter_rows(min_col=1, max_col=10, values_only=True))

        if not result or len(result[0]) != 10:
            mb.showerror("Хатолик", "Файлда 10 та устун бўлиши керак!")
            return

        columns = ['№', 'Ф.И.О', 'Тугилган', 'Жинси', 'Класс', '1-тур', '2-тур', '3-тур', 'Вилоят', 'Кўкрак рақами']
        df = pd.DataFrame(result, columns=columns)

        output_filename = f"Protokol_{sport_turi}.xlsx"
        with pd.ExcelWriter(output_filename, engine='openpyxl') as writer:
            grouped = df.groupby(['Класс', 'Жинси'])

            for (klass, gender), group in grouped:
                filtered = group[
                    (group['1-тур'] == sport_turi) | (group['2-тур'] == sport_turi) | (group['3-тур'] == sport_turi)
                ]
                if filtered.empty:
                    continue

                filtered = filtered[['№', 'Ф.И.О', 'Тугилган', 'Вилоят']]
                sheet_name = self.sanitize_sheet_name(f"{klass}_{gender}")

                # Создаем временный DataFrame с правильной разбивкой
                temp_data = []
                group_number = 1
                chunk_size = 5

                if sport_turi=="Ядро" or sport_turi=='Узунлик' or sport_turi=='Найза' or sport_turi=='Диск' or sport_turi=='Баландлик' or sport_turi=='Клаб':
                    for start in range(0, len(filtered), chunk_size):
                        chunk = filtered.iloc[start:start + chunk_size]
                        temp_data.extend(chunk.values.tolist())
                        print(temp_data)
                else:
                    for start in range(0, len(filtered), chunk_size):
                        chunk = filtered.iloc[start:start + chunk_size]
                        temp_data.append([f"{group_number}-ЮГУРИШ ГУРУХИ", "", "", ""])
                        print(sheet_name,chunk.values.tolist())
                        temp_data.extend(chunk.values.tolist())
                        group_number += 1

                # Добавляем строки судей
                temp_data.extend([
                    ["", "", "", ""],
                    ["","БОШ ХАКАМ:", "", ""],
                    ["", "", "", ""],
                    ["","ХАКАМ:", "", ""],
                    ["", "", "", ""],
                    ["","ХАКАМ:", "", ""],
                    ["", "", "", ""],
                    ["","ХАКАМ:", "", ""]
                ])

                temp_df = pd.DataFrame(temp_data, columns=['№', 'Ф.И.О', 'Тугилган', 'Вилоят'])
                temp_df.to_excel(writer, sheet_name=sheet_name, index=False)

    ##############################
    ## Excelni ochib to'girlash ##
    ##############################
    def obrabotka(self,sport_turi):
        # Загружаем файл
        filename = f"Protokol_{sport_turi}.xlsx"
        #print("Файл существует?", os.path.exists(filename))
        #print("Полный путь:", os.path.abspath(filename))
        wb = load_workbook(filename)

        # Задаем нужный шрифт
        my_font = Font(name='Times New Roman', size=12, bold=False)

        # Проходим по всем листам и всем ячейкам
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    # Проверяем, есть ли значение (чтобы не ставить стиль в пустые)
                    if cell.value is not None:
                        cell.font = my_font


        # Проходим по всем листам
        for ws in wb.worksheets:
            # Добавляем 4 пустые строки сверху
            ws.insert_rows(1, amount=3)

            # Изменяем ширину столбцов (например, от A до I)
            column_widths = {'A':8, 'B': 28, 'C': 13.29, 'D': 8.86, 'E': 15.57, 'F': 9.43, 'G': 8.43, 'H': 10, 'I': 10}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            time_n_r12 = Font(name='Times New Roman', size=12, bold=True)
            if ws.title[-1] == "Э":
                jins = "эркаклар"
            else:
                jins = "аёллар"


            # Sarlavhani to'g'irlash
            ws.merge_cells('A1:G1')
            ws['A1'] = f"Пара енгил атлетика бўйича Ўзбекистон чемпионати\n{ws.title[:-2]} {jins} тоифасида {sport_turi} масофага югуриш мусобақа\nБАЁННОМАСИ"
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # ws['A1'].font = Font(bold=True)
            ws.row_dimensions[1].height = 90
            ws['A1'].font = time_n_r12

            # Путь к картинкам
            left_img_path = "test_images/logoparauz1.png"
            right_img_path = "test_images/laylak.jpg"
            # Добавляем левую картинку
            img_left = Image(left_img_path)
            img_left.width = 90
            img_left.height = 100
            ws.add_image(img_left, "A1")  # В левый верхний угол

            # Добавляем правую картинку
            img_right = Image(right_img_path)
            img_right.width = 80
            img_right.height = 130
            ws.add_image(img_right, "G1")  # В правый верхний угол (или H1, если нужно)

            # Jins satri
            ws.merge_cells('A2:G2')
            ws['A2'] = jins.upper()
            ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A2'].font = time_n_r12
            ws.row_dimensions[2].height = 18

            ws.merge_cells('A3:B3')
            ws['A3'] = "6-8 май 2024 йил"
            ws['A3'].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells('E3:G3')
            ws['E3'] = "Ангрен шаҳри"
            ws['E3'].alignment = Alignment(horizontal="center", vertical="center")


            ws['E4'] = "Кўкрак рақами"
            ws['F4'] = "Натижа"
            ws['G4'] = "Ўрин"

            # 1-yugurish gruhi
            ws.merge_cells('A5:G5')
            ws['A5'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A5'].font = time_n_r12

            # 2-ЮГУРИШ ГУРУХИ
            if ws['A11'].value =="2-ЮГУРИШ ГУРУХИ":
                ws.merge_cells('A11:G11')
                ws['A11'].alignment = Alignment(horizontal="center", vertical="center")
                ws['A11'].font = time_n_r12

            # 3-ЮГУРИШ ГУРУХИ
            if ws['A17'].value =="3-ЮГУРИШ ГУРУХИ":
                ws.merge_cells('A17:G17')
                ws['A17'].alignment = Alignment(horizontal="center", vertical="center")
                ws['A17'].font = time_n_r12

            # Создаём стиль тонкой линии снизу
            thick_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            t = 1  # начинаем с 6-й строки
            while True:
                value = ws[f'B{t}'].value
                if value == "БОШ ХАКАМ:":
                    print(f'Нашёл "БОШ ХАКАМ" в строке {t}')
                    break
                t += 1

            # Familiyalar chizig'i
            thin_bottom = Border(bottom=Side(style='thin', color='000000'))
            ws[f'E{t}'].border = thin_bottom
            ws[f'F{t}'].border = thin_bottom
            ws[f'G{t}'].border = thin_bottom
            ws[f'E{t+2}'].border = thin_bottom
            ws[f'F{t+2}'].border = thin_bottom
            ws[f'G{t+2}'].border = thin_bottom

            ws[f'E{t+4}'].border = thin_bottom
            ws[f'F{t+4}'].border = thin_bottom
            ws[f'G{t+4}'].border = thin_bottom

            ws[f'E{t+6}'].border = thin_bottom
            ws[f'F{t+6}'].border = thin_bottom
            ws[f'G{t+6}'].border = thin_bottom

            # Jadval chizig'i
            for rr in range(4,t-1):
                for col in range(1, 8):  # от столбца A (1) до G (7)
                    cell = ws.cell(row=rr, column=col)
                    cell.border = thick_border

        # Сохраняем результат в новый файл
        try:
            wb.save(f"Protokolob_{sport_turi}.xlsx")
        except Exception:
            print("Bunday nomli fayl hozir ochiq undan chiqib ketishingizni so'rayman!!!")


    # Ikkinchi jadvalning obrabotkasi
    def obrabotka_2jadval(self,sport_turi):
        # Загружаем файл
        filename = f"Protokol_{sport_turi}.xlsx"
        #print("Файл существует?", os.path.exists(filename))
        #print("Полный путь:", os.path.abspath(filename))
        wb = load_workbook(filename)

        # Задаем нужный шрифт
        my_font = Font(name='Times New Roman', size=12, bold=False)

        # Проходим по всем листам и всем ячейкам
        for ws in wb.worksheets:
            for row in ws.iter_rows():
                for cell in row:
                    # Проверяем, есть ли значение (чтобы не ставить стиль в пустые)
                    if cell.value is not None:
                        cell.font = my_font


        # Проходим по всем листам
        for ws in wb.worksheets:
            # Добавляем 4 пустые строки сверху
            ws.insert_rows(1, amount=3)

            # Изменяем ширину столбцов (например, от A до I)
            column_widths = {'A':6, 'B': 23, 'C': 13, 'D': 8.86, 'E': 10, 'F': 9, 'G': 11, 'H': 11, 'I': 11,'J':11,'K':11,'L':11,'M':11,'N':11}
            for col, width in column_widths.items():
                ws.column_dimensions[col].width = width
            time_n_r12 = Font(name='Times New Roman', size=12, bold=True)
            if ws.title[-1] == "Э":
                jins = "эркаклар"
            else:
                jins = "аёллар"


            # Sarlavhani to'g'irlash
            # sport_turlari=('Ядро','100м','1500м','200м','400м','Баландлик','Диск','Клаб','Найза','Узунлик','Арава')
            if sport_turi=="Ядро":
                tt=f"""Пара енгил атлетика бўйича Ўзбекистон чемпионати\nЯдро улоқтириш бўйича {ws.title[:-2]} тоифасидаги\nспортчилар учун мусобақа\nБАЁННОМАСИ"""
            elif sport_turi=="Клаб":
                tt=f"""Пара енгил атлетика бўйича Ўзбекистон чемпионати\nCLUB THROW улоқтириш бўйича {ws.title[:-2]} тоифасидаги\nспортчилар учун мусобақа\nБАЁННОМАСИ"""
            elif sport_turi=="Баландлик":
                tt=f"""Пара енгил атлетика бўйича Ўзбекистон чемпионати\nБаландликка сакраш бўйича {ws.title[:-2]} тоифасидаги\nспортчилар учун мусобақа\nБАЁННОМАСИ"""
            elif sport_turi=="Диск":
                tt=f"""Пара енгил атлетика бўйича Ўзбекистон чемпионати\nДиск улоқтириш бўйича {ws.title[:-2]} тоифасидаги\nспортчилар учун мусобақа\nБАЁННОМАСИ"""
            elif sport_turi=="Найза":
                tt=f"""Пара енгил атлетика бўйича Ўзбекистон чемпионати\nЯдро улоқтириш бўйича {ws.title[:-2]} тоифасидаги\nспортчилар учун мусобақа\nБАЁННОМАСИ"""
            elif sport_turi=="Узунлик":
                tt=f"""Пара енгил атлетика бўйича Ўзбекистон чемпионати\nУзунликка сакраш бўйича {ws.title[:-2]} тоифасидаги\nспортчилар учун мусобақа\nБАЁННОМАСИ"""
            else:
                tt="*-*-*-*"

            ws.merge_cells('A1:N1')
            ws['A1'] = tt
            ws['A1'].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            # ws['A1'].font = Font(bold=True)
            ws.row_dimensions[1].height = 90
            ws['A1'].font = time_n_r12

            # Путь к картинкам
            left_img_path = "test_images/logoparauz1.png"
            right_img_path = "test_images/laylak.jpg"
            # Добавляем левую картинку
            img_left = Image(left_img_path)
            img_left.width = 90
            img_left.height = 100
            ws.add_image(img_left, "A1")  # В левый верхний угол

            # Добавляем правую картинку
            img_right = Image(right_img_path)
            img_right.width = 80
            img_right.height = 130
            ws.add_image(img_right, "N1")  # В правый верхний угол (или H1, если нужно)

            # Jins satri
            ws.merge_cells('A2:N2')
            ws['A2'] = jins.upper()
            ws['A2'].alignment = Alignment(horizontal="center", vertical="center")
            ws['A2'].font = time_n_r12
            ws.row_dimensions[2].height = 18

            ws.merge_cells('A3:B3')
            ws['A3'] = "6-8 май 2024 йил"
            ws['A3'].font = time_n_r12
            ws['A3'].alignment = Alignment(horizontal="center", vertical="center")

            ws.merge_cells('K3:N3')
            ws['K3'] = "Ангрен шаҳри"
            ws['K3'].font = time_n_r12
            ws['K3'].alignment = Alignment(horizontal="center", vertical="center")

            ws['E4'] = "Класс"
            ws['F4'] = "К/К рақми"
            ws['G4'] = "1-уриниш"
            ws['H4'] = "2-уриниш"
            ws['I4'] = "3-уриниш"
            ws['J4'] = "финалга ўтиш"
            ws['K4'] = "4-уриниш"
            ws['L4'] = "5-уриниш"
            ws['M4'] = "6-уриниш"
            ws['N4'] = "ўрни"

            # Jadval shapkasini qoraytirish
            for col in range(1, 15):  # от столбца A (1) до n (14)
                cell = ws.cell(row=4, column=col)
                cell.font = time_n_r12
                cell.alignment = Alignment(horizontal="center", vertical="center")


            # Создаём стиль тонкой линии снизу
            thick_border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            t = 1  # начинаем с 6-й строки
            while True:
                value = ws[f'B{t}'].value
                if value == "БОШ ХАКАМ:":
                    #print(f'Нашёл "БОШ ХАКАМ" в строке {t}')
                    break
                t += 1

            # Familiyalar chizig'i
            thin_bottom = Border(bottom=Side(style='thin', color='000000'))
            ws[f'E{t}'].border = thin_bottom
            ws[f'F{t}'].border = thin_bottom
            ws[f'G{t}'].border = thin_bottom
            ws[f'E{t+2}'].border = thin_bottom
            ws[f'F{t+2}'].border = thin_bottom
            ws[f'G{t+2}'].border = thin_bottom

            ws[f'E{t+4}'].border = thin_bottom
            ws[f'F{t+4}'].border = thin_bottom
            ws[f'G{t+4}'].border = thin_bottom

            ws[f'E{t+6}'].border = thin_bottom
            ws[f'F{t+6}'].border = thin_bottom
            ws[f'G{t+6}'].border = thin_bottom

            # Jadval chizig'i
            for rr in range(4,t-1):
                for col in range(1, 15):  # от столбца A (1) до G (7)
                    cell = ws.cell(row=rr, column=col)
                    cell.border = thick_border

        # Сохраняем результат в новый файл
        try:
            wb.save(f"Protokolob_{sport_turi}.xlsx")
        except Exception:
            print("Bunday nomli fayl hozir ochiq undan chiqib ketishingizni so'rayman!!!")



# sport="1500м"
# sport_turlari_1jadval=('100м','1500м','200м','400м','Арава')
# sport_turlari_2jadval=('Ядро','Баландлик','Диск','Клаб','Найза','Узунлик')
# # exeldan olib oddiy jadval hosil qilish
# exeldanolish(sport)
# if sport in sport_turlari_1jadval:
#     # obrabotka_1_jadval
#     obrabotka(sport)
# elif sport in sport_turlari_2jadval:
#     # obrabotka_2_jadval
#     obrabotka_2jadval(sport)